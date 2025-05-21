import openpyxl
import unicodedata
from pathlib import Path
from typing import List, Dict, Any

def process_str(s: str) -> str:
    """
    Remove espaços antes/depois, converte para lowercase, remove acentos
    e substitui espaços internos por underscores.
    """
    raw = str(s or "").strip()
    normalized = unicodedata.normalize("NFD", raw)
    no_accents = "".join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    return no_accents.lower().replace(" ", "_")


def extract(filepath: str, headers: str, firstRow: int, lastRow: int) -> List[List[str]]:
    """
    Lê os dados para cada linha de firstRow a lastRow
    retorna lista de disciplinas reprovadas para cada aluno.
    """
    path = Path(filepath)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    cols = [h.strip() for h in headers.split(",") if h.strip()]
    header_map: Dict[str, str] = {
        col: process_str(ws[f"{col}1"].value)
        for col in cols
    }
    results: List[List[str]] = []
    for i in range(firstRow, lastRow + 1):
        subjects: List[str] = []
        for col in cols:
            cell = ws[f"{col}{i}"]
            raw = cell.value
            score = float(raw) if isinstance(raw, (int, float)) else 0.0
            if score < 6.0:
                subjects.append(header_map[col])
        if subjects:
            results.append(subjects)
    return results


def extract_json(filepath: str, headers: str, firstRow: int, lastRow: int) -> Dict[str, Any]:
    """
    Gera JSON com nome da planilha (sem .xlsx) como chave,
    e índice->lista de disciplinas reprovadas.
    """
    path = Path(filepath)
    sheet_name = path.stem
    data = extract(filepath, headers, firstRow, lastRow)
    mapped: Dict[str, Any] = {}
    for idx, subjects in enumerate(data):
        mapped[str(idx)] = subjects
    return {sheet_name: mapped}

def merge_jsons(json_list: List[Dict[str, Any]]) -> Dict[str, Any]:
    merged: Dict[str, Any] = {}
    for jd in json_list:
        for key, value in jd.items():
            if key in merged:
                merged[key].update(value)
            else:
                merged[key] = value.copy()
    return merged